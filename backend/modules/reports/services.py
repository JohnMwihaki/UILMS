import csv
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_csv_report(filename, headers, rows):
    """
    Generates a CSV file response.
    """
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
        
    return response

def generate_excel_report(sheet_name, headers, rows):
    """
    Generates a styled Excel spreadsheet response.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.views.sheetView[0].showGridLines = True

    # Font styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")

    # Fills & borders
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Title row
    ws.append([f"Karatina University - UILMS Report: {sheet_name}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = title_font
    ws.row_dimensions[1].height = 25
    ws.append([]) # Spacer row

    # Headers
    ws.append(headers)
    ws.row_dimensions[3].height = 20
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data Rows
    for row_data in rows:
        ws.append(row_data)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 18
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Skip first row because it is merged title
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Output Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{sheet_name.lower().replace(" ", "_")}.xlsx"'
    
    # Save workbook to memory and output
    buffer = io.BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response

def generate_pdf_report(title_text, headers, rows):
    """
    Generates a professionally styled PDF report using ReportLab.
    """
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title_text.lower().replace(" ", "_")}.pdf"'

    # We use landscape to fit tables nicely
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1F497D"),
        spaceAfter=15,
        alignment=1 # Center
    )

    header_style = ParagraphStyle(
        name="HeaderStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )

    cell_style = ParagraphStyle(
        name="CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10
    )

    elements = []
    
    # Header Info
    elements.append(Paragraph("KARATINA UNIVERSITY", title_style))
    elements.append(Paragraph(f"Department of Biological and Physical Sciences — UILMS Report: {title_text}", ParagraphStyle(
        name="SubTitle",
        parent=styles["Normal"],
        fontSize=10,
        alignment=1,
        spaceAfter=20
    )))

    # Table Formatting
    formatted_headers = [Paragraph(h, header_style) for h in headers]
    formatted_rows = []
    for r in rows:
        formatted_rows.append([Paragraph(str(cell) if cell is not None else "", cell_style) for cell in r])

    table_data = [formatted_headers] + formatted_rows

    # Calculate column widths
    # Standard letter width is 612, landscape letter is 792.
    # Margins are 30 each side, so printable area width is 732.
    col_width = 732 / len(headers)
    
    table = Table(table_data, colWidths=[col_width] * len(headers))
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F497D")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D3D3")),
        # Alternating row backgrounds
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)
    
    return response
